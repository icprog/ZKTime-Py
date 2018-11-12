#!/usr/bin/python3
# coding=utf-8


"""
考勤机脚本入口
"""

__author__ = 'LCD'

import sys
import win32com.client
from rx.subjects import Subject
from cd_tools import *
import json
import openpyxl
import xlwt
import requests

zk = win32com.client.Dispatch('zkemkeeper.ZKEM.1')
key_am = 'am'
key_pm = 'pm'
key_first = 'first'
key_last = 'last'
key_name = 'name'
key_record = 'record'
'''
    log_data = {'userId': {'yy/mm/dd': {'am': {'first': '',
                                               'last': '',
                                               'name': ''},
                                        'pm': {'first': '',
                                               'last': '',
                                               'name': ''}}}}

'''
post_dates = []
last_timestamp = 0.0


def user_name_to_gbk(name):
    return str(name.split(u'\x00')[0].encode('gbk'), encoding="gbk")


def save_json(obj):
    d = json.dumps(obj, ensure_ascii=False)
    file = 'Erazar考勤记录.json'
    json_file = open(file, 'w')
    json_file.write(d)
    json_file.close()


def save_date_json(obj):
    d = json.dumps(obj, ensure_ascii=False)
    file = 'Erazar上次读取日期和时间记录.json'
    json_file = open(file, 'w')
    json_file.write(d)
    json_file.close()


def open_json_file():
    # print('open_json_file')
    try:
        json_file = open("Erazar上次读取时间记录.json", 'r')
    except:
        make_date_begin(cd_make_date_for_now(-7))
    else:
        read_json(json_file)


def read_json(json_file):
    try:
        dic = json.load(json_file)
    except:
        read_dict({'date': cd_make_date_for_now(-7), 'timestamp': ''})
    else:
        read_dict(dic)


def read_dict(dic):
    try:
        date = dic['date']
    except:
        make_date_begin(cd_make_date_for_now(-7))
    else:
        make_date_begin(date)

    try:
        timestamp = dic['timestamp']
    except:
        make_last_timestamp()
    else:
        make_last_timestamp(timestamp)


def make_date_begin(last):
    global post_dates
    print('---------->上一次时间')
    print(last)
    post_dates = cd_make_date_list(last, cd_time_now('%Y-%m-%d'), '/')
    print('---------->时间组')
    print(post_dates)
    post_dates.append(cd_join_change(cd_time_now('%Y-%m-%d'), '-', '/'))

def make_date_begin_request(star, end):
    global post_dates
    last = cd_timestamp_to_time(star, '%Y-%m-%d')
    now = cd_timestamp_to_time(end, '%Y-%m-%d')
    post_dates = cd_make_date_list(last, now, '/')
    post_dates.append(cd_join_change(now, '-', '/'))

def make_last_timestamp(last=0.0):
    global last_timestamp
    print('---------->上一次时间戳')
    print(last)
    last_timestamp = last


class ReadUserData(object):
    # 用户Id
    key_id = 'userId'
    # 用户名
    key_name = 'userName'
    # 当天首次打卡时间
    key_first = 'userFirst'
    # 当天最后一次打卡时间
    key_last = 'userLast'
    
    start_time = 0
    end_time = 0

    def __init__(self, type=0):
        self.output_type = type
        self.post_num = 0
        self.user = {}
        self.log_data = {}
        self.source = Subject()
        self.__connect_net()

    def __connect_net(self):
        print('-----连接考勤机----->')
        if not zk.Connect_Net('192.168.1.238', 4370):
            print('-----连接失败----->')
            # sys.exit(1)
            self.source.on_error('连接失败')

    def read_user_all(self):
        print('-----读取所有员工----->')
        while 1:
            zk_num, user_id, user_name, pwd, privilege, state = zk.SSR_GetAllUserInfo(1)
            # print(user_id + user_name + str(state))
            if not zk_num:
                # print('break')
                break
            elif state:
                self.user[user_id] = user_name_to_gbk(user_name)
        # print(self.user)

    def make_log_data(self):
        if self.output_type==0:
            self.make_excel_json()
        elif self.output_type == 1:
            self.make_post_json()
        else:
            self.make_post_json_all()

    def make_post_json(self):
        global post_dates
        # dates = cd_make_date_list(cd_make_date_for_now(-1), cd_time_now('%Y-%m-%d'), '/')
        for item in self.user:
            self.log_data[item] = {}
            for t in post_dates:
                self.log_data[item][t] = {key_name: self.user[item],
                                          key_first: '',
                                          key_last: ''
                                          }

    def make_post_json_all(self):
        global post_dates
        # dates = cd_make_date_list(cd_make_date_for_now(-1), cd_time_now('%Y-%m-%d'), '/')
        for item in self.user:
            self.log_data[item] = {}
            for t in post_dates:
                self.log_data[item][t] = {key_name: self.user[item],
                                          key_record: []
                                          }

    def make_excel_json(self):
        global post_dates
        # dates = cd_make_date_list(cd_make_date_for_now(-29), cd_time_now('%Y-%m-%d'), '/')
        for item in self.user:
            self.log_data[item] = {}
            for t in post_dates:
                self.log_data[item][t] = {key_am: {key_name: self.user[item],
                                                   key_first: '',
                                                   key_last: ''},
                                          key_pm: {key_name: self.user[item],
                                                   key_first: '',
                                                   key_last: ''}
                                          }

    def read_user_for_id(self, user_id):
        while 1:
            zk_num, user_id, user_name, pwd, privilege, state = zk.SSR_GetAllUserInfo(1)
            if not zk_num:
                # print('break')
                break
            elif state:
                return user_name_to_gbk(user_name)

    def prefetch_all_log_data(self):
        print('-----读取考勤记录----->')
        if zk.ReadAllGLogData(1):  # read All checkin data
            self.read_all_log_data()
        else:
            print('-----读取考勤记录失败----->')

    def read_all_log_data(self):
        """
        :return: 读取考勤记录之前先记录当前时间节点
        """
        save_date_json({'date': cd_time_now('%Y-%m-%d'), 'timestamp': cd_timestamp()})
        while 1:
            zk_num, user_id, verify, state, yy, month, dd, hh, mm, ss, code = zk.SSR_GetGeneralLogData(1)
            if not zk_num:
                # print('break--1')
                break
            if self.output_type == 0:
                self.data_processing_excel_json(user_id, yy, month, dd, hh, mm)
            elif self.output_type == 1:
                self.data_processing_post_json(user_id, yy, month, dd, hh, mm)
            else:
                self.data_processing_post_json_all(user_id, yy, month, dd, hh, mm)

    def time_processing(self, s):
        s = str(s)
        if len(s) == 1:
            s = '0' + s
        return s

    def data_processing_post_json(self, user_id, yy, month, dd, hh, mm):
        """
        :param user_id:用户ID
        :param yy:
        :param month:
        :param dd:
        :param hh:
        :param mm:
        :return:
        1, 检查当前是否有该用户
        2，检查该用户当前是否已打卡，写入打卡时间，没有则存入first，有则存入last
        获取时间段
        以时间段 和用户 id 建立  json
        """
        month = self.time_processing(month)
        dd = self.time_processing(dd)
        hh = self.time_processing(hh)
        mm = self.time_processing(mm)
        value_date = str(yy) + '/' + month + '/' + dd
        value_time = hh + ':' + mm
        if user_id in self.log_data:
            if value_date in self.log_data[user_id]:
                # print('记录---------->')
                f = self.log_data[user_id][value_date][key_first]
                l = self.log_data[user_id][value_date][key_last]
                if len(f) == 0:
                    # print('上班：' + user_id + '-->' + value_date + '---' + value_time)
                    self.log_data[user_id][value_date][key_first] = value_time
                    self.log_data[user_id][value_date][key_last] = value_time
                else:
                    self.log_data[user_id][value_date][key_last] = value_time

    def data_processing_post_json_all(self, user_id, yy, month, dd, hh, mm):
        """
        :param user_id:用户ID
        :param yy:
        :param month:
        :param dd:
        :param hh:
        :param mm:
        :return:
        1, 检查当前是否有该用户
        2，检查该用户当前是否已打卡，写入打卡时间，没有则存入first，有则存入last
        获取时间段
        以时间段 和用户 id 建立  json
        """
        month = self.time_processing(month)
        dd = self.time_processing(dd)
        hh = self.time_processing(hh)
        mm = self.time_processing(mm)
        value_date = str(yy) + '/' + month + '/' + dd
        value_time = hh + ':' + mm
        value_timestamp = cd_time_to_timestamp(str(yy)+'-'+month+'-'+dd+' '+hh+':'+mm+':00')
        if (value_timestamp > self.start_time) and (value_timestamp <= self.end_time):
            if user_id in self.log_data:
                if value_date in self.log_data[user_id]:
                    # print('记录---------->')
                    self.log_data[user_id][value_date][key_record].append(value_time)

    def data_processing_excel_json(self, user_id, yy, month, dd, hh, mm):
        """
        :param user_id:用户ID
        :param yy:
        :param month:
        :param dd:
        :param hh:
        :param mm:
        :return:
        1, 检查当前是否有该用户
        2，检查该用户当前是否已打卡，写入打卡时间，没有则存入first，有则存入last
        获取时间段
        以时间段 和用户 id 建立  json

        """
        month = self.time_processing(month)
        dd = self.time_processing(dd)
        hh = self.time_processing(hh)
        mm = self.time_processing(mm)
        value_date = str(yy) + '/' + month + '/' + dd
        value_time = hh + ':' + mm
        # print('记录：' + user_id + '-->' + value_date + '---' + value_time)

        if user_id in self.log_data:
            if value_date in self.log_data[user_id]:
                # print('记录---------->')
                am_f = self.log_data[user_id][value_date][key_am][key_first]
                am_l = self.log_data[user_id][value_date][key_am][key_last]
                pm_f = self.log_data[user_id][value_date][key_pm][key_first]
                pm_l = self.log_data[user_id][value_date][key_pm][key_last]
                if len(am_f) == 0 and value_time <= '13:30':
                    # print('上班：' + user_id + '-->' + value_date + '---' + value_time)
                    self.log_data[user_id][value_date][key_am][key_first] = value_time
                    self.log_data[user_id][value_date][key_am][key_last] = value_time
                elif value_time <= '13:30':
                    # print('上午下班： ' + user_id + '-->' + value_date + '---' + value_time)
                    self.log_data[user_id][value_date][key_am][key_last] = value_time
                elif len(pm_f) == 0 and value_time > '13:30':
                    self.log_data[user_id][value_date][key_pm][key_first] = value_time
                    self.log_data[user_id][value_date][key_pm][key_last] = value_time
                else:
                    self.log_data[user_id][value_date][key_pm][key_last] = value_time
        """
        if user_id in self.log_data:
            if value_date in self.log_data[user_id]:
                # print('记录---------->')
                am_f = self.log_data[user_id][value_date][key_am][key_first]
                am_l = self.log_data[user_id][value_date][key_am][key_last]
                pm_f = self.log_data[user_id][value_date][key_pm][key_first]
                pm_l = self.log_data[user_id][value_date][key_pm][key_last]
                if len(am_f) == 0 and len(pm_f) == 0:
                    if value_time <= '12:00':
                        # print('上班：' + user_id + '-->' + value_date + '---' + value_time)
                        self.log_data[user_id][value_date][key_am][key_first] = value_time
                        self.log_data[user_id][value_date][key_am][key_last] = value_time
                    else:
                        self.log_data[user_id][value_date][key_pm][key_first] = value_time
                        self.log_data[user_id][value_date][key_pm][key_last] = value_time
                else:
                    # print('下班：' + user_id + '-->' + value_date + '---' + value_time)
                    if value_time <= '12:00':
                        self.log_data[user_id][value_date][key_am][key_last] = value_time
                    else:
                        self.log_data[user_id][value_date][key_am][key_last] = ''
                        self.log_data[user_id][value_date][key_pm][key_last] = value_time
        """

    def save_excel_xlsx(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        i = 2
        self.make_sheet_top_xlsx(ws)
        for u_id in self.log_data.keys():
            # 取id
            for t in self.log_data[u_id].keys():
                # 取日期
                for m in self.log_data[u_id][t].keys():
                    # 取上下午
                    ws.cell(row=i, column=1, value=str(i - 1))
                    ws.cell(row=i, column=2, value=u_id)
                    ws.cell(row=i, column=3, value=u_id)
                    ws.cell(row=i, column=4, value=self.log_data[u_id][t][m][key_name])
                    ws.cell(row=i, column=6, value=t)
                    if m == key_am:
                        ws.cell(row=i, column=7, value='上午')
                        ws.cell(row=i, column=8, value='09:00')
                        ws.cell(row=i, column=9, value='12:00')
                    else:
                        ws.cell(row=i, column=7, value='下午')
                        ws.cell(row=i, column=8, value='13:30')
                        ws.cell(row=i, column=9, value='23:00')
                    ws.cell(row=i, column=10, value=self.log_data[u_id][t][m][key_first])
                    ws.cell(row=i, column=11, value=self.log_data[u_id][t][m][key_last])
                    i += 1
        wb.save('考勤记录.xlsx')
        print('----------> 完成 考勤记录.xlsx')

    def make_sheet_top_xlsx(self, ws):
        ws.cell(row=1, column=1, value='序号')
        ws.cell(row=1, column=2, value='考勤号码')
        ws.cell(row=1, column=3, value='自定义编号')
        ws.cell(row=1, column=4, value='姓名')
        ws.cell(row=1, column=5, value='是否智能排班')
        ws.cell(row=1, column=6, value='日期')
        ws.cell(row=1, column=7, value='对应时段')
        ws.cell(row=1, column=8, value='上班时间')
        ws.cell(row=1, column=9, value='下班时间')
        ws.cell(row=1, column=10, value='签到时间')
        ws.cell(row=1, column=11, value='签退时间')
        ws.cell(row=1, column=12, value='应到')
        ws.cell(row=1, column=13, value='实到')
        ws.cell(row=1, column=14, value='迟到时间')
        ws.cell(row=1, column=15, value='早退时间')
        ws.cell(row=1, column=16, value='是否旷工')
        ws.cell(row=1, column=17, value='加班时间')
        ws.cell(row=1, column=18, value='工作时间')
        ws.cell(row=1, column=19, value='例外情况')
        ws.cell(row=1, column=20, value='应签到')
        ws.cell(row=1, column=21, value='应签退')
        ws.cell(row=1, column=22, value='部门')
        ws.cell(row=1, column=23, value='平日')
        ws.cell(row=1, column=24, value='周末')
        ws.cell(row=1, column=25, value='节假日')
        ws.cell(row=1, column=26, value='出勤时间')
        ws.cell(row=1, column=27, value='平日加班')
        ws.cell(row=1, column=28, value='周末加班')
        ws.cell(row=1, column=29, value='节假日加班')

    def save_excel_xls(self):
        wb = xlwt.Workbook(encoding='utf-8', style_compression=0)
        ws = wb.add_sheet('sheet', cell_overwrite_ok=True)
        # 直接写更直观，不用循环
        ws.write(0, 0, '序号')
        ws.write(0, 1, '考勤号码')
        ws.write(0, 2, '自定义编号')
        ws.write(0, 3, '姓名')
        ws.write(0, 4, '是否智能排班')
        ws.write(0, 5, '日期')
        ws.write(0, 6, '对应时段')
        ws.write(0, 7, '上班时间')
        ws.write(0, 8, '下班时间')
        ws.write(0, 9, '签到时间')
        ws.write(0, 10, '签退时间')
        i = 1
        for u_id in self.log_data.keys():
            # 取id
            for t in self.log_data[u_id].keys():
                # 取日期
                for m in self.log_data[u_id][t].keys():
                    # 取上下午
                    ws.write(i, 0, str(i - 1))
                    ws.write(i, 1, u_id)
                    ws.write(i, 2, u_id)
                    ws.write(i, 3, self.log_data[u_id][t][m][key_name])
                    ws.write(i, 5, t)
                    if m == key_am:
                        ws.write(i, 6, '上午')
                        ws.write(i, 7, '09:00')
                        ws.write(i, 8, '12:00')
                    else:
                        ws.write(i, 6, '下午')
                        ws.write(i, 7, '13:30')
                        ws.write(i, 8, '23:00')
                    ws.write(i, 9, self.log_data[u_id][t][m][key_first])
                    ws.write(i, 10, self.log_data[u_id][t][m][key_last])
                    i += 1
        xls = cd_time_now('%Y-%m-%d %H-%M') + '.xls'
        wb.save(xls)
        print('----------> 完成 输出 ' + xls)

    def post_json_to(self):
        if self.post_num > 10:
            return
        url = 'https://xxxxxxxx.com'
        body = {'data': self.log_data}
        try:
            r = requests.post(url, data=body)
        except:
            self.post_num += 1
            self.post_json_to()
        else:
            res = r.json()
            code = res['code']
            # print(res)
            if code == 200:
                self.login_ok(res)
            else:
                self.source.on_error(FileNotFoundError)


if __name__ == '__main__':

    t = ReadUserData(2)
    t.read_user_all()
    t.make_log_data()
    t.prefetch_all_log_data()
    # t.save_excel_xlsx()
    # t.save_excel_xls()
    save_json(t.log_data)
    # save_date_json({'date': cd_time_now('%Y-%m-%d'), 'timestamp': cd_timestamp()})
    # print(json.dumps(t.log_data))
